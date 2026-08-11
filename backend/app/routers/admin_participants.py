import io
import csv
import json
from datetime import datetime, timezone
from typing import List, Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, UploadFile, File, Form, status
from openpyxl import load_workbook

from app.core.deps import get_current_admin_user
from app.db.supabase_client import supabase
from app.models.participant_schemas import ImportUploadResponse, ImportDetail, ImportActionResponse
from app.services.document_parser import extract_text
from app.services.summarizer import client as anthropic_client  # reuses the existing Anthropic client setup
from app.services.geocoding import geocode_district

router = APIRouter(prefix="/admin/participants", tags=["admin-participants"])

MAX_FILE_SIZE = 25 * 1024 * 1024
ALLOWED_EXTENSIONS = {"xlsx", "csv", "pdf", "docx"}

HEADER_ALIASES = {
    "hh_size": "household_size",
    "per_capita_baseline_ugx": "pre_program_income",
    "breadwinner": "main_breadwinner",
    "age": "age",
    "education": "highest_education",
    "pre_emp_status": "employed_before",
    "pre_job_title": "employed_before_type",
    "district": "district",
    "country_of_birth": "country",
    "household_size": "household_size",
    "pre_program_income": "pre_program_income",
    "main_breadwinner": "main_breadwinner",
    "highest_education": "highest_education",
    "employed_before": "employed_before",
    "employed_before_type": "employed_before_type",
    "country": "country",
}

DISTRICT_ALIASES = {
    "Namugongo- Kyaliwajjala": "Kyaliwajjala",
    "Ndejje, Namasuba": "Namasuba",
    "Rubanda, Kabale": "Rubanda",
    "Fortportal": "Fort Portal",
    "Kiryandogo": "Kiryandongo",
}

EXPECTED_COLUMNS = [
    "household_size", "pre_program_income", "main_breadwinner",
    "age", "highest_education", "employed_before",
    "employed_before_type", "district", "country",
    "graduation_status",
]

INTEGER_FIELDS = {"household_size", "age"}
NUMERIC_FIELDS = {"pre_program_income"}
BOOLEAN_FIELDS = {"employed_before"}

UGX_TO_USD = 1 / 3750


def _coerce_value(field: str, value):
    if value is None or value == "":
        return None

    if field == "pre_program_income":
        try:
            return round(float(value) * UGX_TO_USD, 2)
        except (ValueError, TypeError):
            return None

    if field in INTEGER_FIELDS:
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    if field in NUMERIC_FIELDS:
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    if field in BOOLEAN_FIELDS:
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        if "unemploy" in text:
            return False
        if "employ" in text:
            return True
        return None

    return value


def _map_row(raw_row: dict) -> dict:
    mapped = {}
    for raw_key, value in raw_row.items():
        internal_field = HEADER_ALIASES.get(raw_key)
        if internal_field:
            mapped[internal_field] = _coerce_value(internal_field, value)
    return {k: mapped.get(k) for k in EXPECTED_COLUMNS}


def _parse_spreadsheet_rows(filename: str, file_bytes: bytes) -> List[dict]:
    ext = filename.rsplit(".", 1)[-1].lower()
    rows: List[dict] = []

    if ext == "xlsx":
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = workbook.active
        header = [str(c).strip().lower().replace(" ", "_") if c else "" for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            raw_row = {header[i]: row[i] for i in range(min(len(header), len(row)))}
            rows.append(_map_row(raw_row))

    elif ext == "csv":
        text = file_bytes.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            raw_row = {k.strip().lower().replace(" ", "_"): v for k, v in row.items()}
            rows.append(_map_row(raw_row))

    for row in rows:
        if not row.get("graduation_status"):
            row["graduation_status"] = "enrolled"

    return rows


def _parse_via_ai(filename: str, file_bytes: bytes) -> List[dict]:
    """
    PDF/DOCX — text extracted via the existing document_parser, then an AI
    step attempts to structure it into rows matching EXPECTED_COLUMNS.
    Real accuracy risk, per the contract — unlike Excel/CSV this is AI
    inference, not exact extraction. Treat output as needing closer admin
    review than spreadsheet-sourced data. Unverified until tested against
    a real PDF/DOCX participant file.
    """
    raw_text = extract_text(filename, file_bytes)

    prompt = f"""Extract participant records from the text below into a JSON array.
Each object must have exactly these fields (use null if not found):
household_size (number), pre_program_income (number, in UGX if given),
main_breadwinner (string), age (number), highest_education (string),
employed_before (true/false), employed_before_type (string),
district (string), country (string).

Return ONLY the JSON array, no other text.

TEXT:
{raw_text[:15000]}
"""

    response = anthropic_client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        messages=[{"role": "user", "content": prompt}],
    )

    raw_output = response.content[0].text.strip()
    if raw_output.startswith("```"):
        raw_output = raw_output.strip("`").lstrip("json").strip()

    parsed = json.loads(raw_output)
    return [{k: _coerce_value(k, row.get(k)) for k in EXPECTED_COLUMNS} for row in parsed]


def _process_import(import_id: str, filename: str, file_bytes: bytes):
    ext = filename.rsplit(".", 1)[-1].lower()

    try:
        if ext in ("xlsx", "csv"):
            rows = _parse_spreadsheet_rows(filename, file_bytes)
        else:
            rows = _parse_via_ai(filename, file_bytes)

        preview_data = {
            "columns": EXPECTED_COLUMNS,
            "sample_rows": rows[:10],
            "all_rows": rows,
            "ai_assisted": ext in ("pdf", "docx"),
        }

        supabase.table("participant_imports").update({
            "status": "pending_review",
            "row_count": len(rows),
            "preview_data": preview_data,
        }).eq("id", import_id).execute()

    except Exception as e:
        supabase.table("participant_imports").update({
            "status": "failed",
            "preview_data": {"error": str(e)},
        }).eq("id", import_id).execute()


@router.post("/import", response_model=ImportUploadResponse, status_code=status.HTTP_202_ACCEPTED)
async def import_participants(
    background_tasks: BackgroundTasks,
    cohort_id: str = Form(...),
    file: UploadFile = File(...),
    admin: dict = Depends(get_current_admin_user),
):
    cohort_check = supabase.table("cohorts").select("id").eq("id", cohort_id).execute()
    if not cohort_check.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Cohort not found.")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="File must be Excel, CSV, PDF, or DOCX.")

    file_bytes = await file.read()
    if len(file_bytes) > MAX_FILE_SIZE:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="File must be under 25MB.")

    result = (
        supabase.table("participant_imports")
        .insert({
            "filename": file.filename,
            "file_type": ext,
            "status": "processing",
            "uploaded_by": admin["id"],
            "cohort_id": cohort_id,
        })
        .execute()
    )
    record = result.data[0]

    background_tasks.add_task(_process_import, record["id"], file.filename, file_bytes)

    return ImportUploadResponse(id=record["id"], filename=record["filename"], status=record["status"])


@router.get("/imports", response_model=List[ImportDetail])
async def list_imports(admin: dict = Depends(get_current_admin_user)):
    result = supabase.table("participant_imports").select(
        "id, filename, file_type, status, row_count, uploaded_by, created_at, confirmed_at"
    ).order("created_at", desc=True).execute()
    return result.data


@router.get("/imports/{import_id}", response_model=ImportDetail)
async def get_import(import_id: str, admin: dict = Depends(get_current_admin_user)):
    result = supabase.table("participant_imports").select("*").eq("id", import_id).execute()
    if not result.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found.")
    return result.data[0]


@router.post("/imports/{import_id}/confirm", response_model=ImportActionResponse)
async def confirm_import(import_id: str, admin: dict = Depends(get_current_admin_user)):
    existing = supabase.table("participant_imports").select("*").eq("id", import_id).execute()
    if not existing.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found.")

    record = existing.data[0]
    if record["status"] != "pending_review":
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Only imports in 'pending_review' status can be confirmed.")

    # Use every parsed row, not just the 10-row preview sample.
    rows = record.get("preview_data", {}).get("all_rows") or record.get("preview_data", {}).get("sample_rows", [])
    for row in rows:
        clean_row = dict(row)
        if clean_row.get("district") in DISTRICT_ALIASES:
            clean_row["district"] = DISTRICT_ALIASES[clean_row["district"]]
        clean_row["source_import_id"] = import_id
        clean_row["cohort_id"] = record.get("cohort_id")
        supabase.table("participants").insert(clean_row).execute()

    # Auto-geocode any Uganda district we haven't seen before, so future
    # imports with brand-new districts get map coordinates automatically,
    # with no manual lookup required.
    unique_districts = set(
        row.get("district") for row in rows
        if row.get("country") == "Uganda" and row.get("district")
    )
    for district in unique_districts:
        district = DISTRICT_ALIASES.get(district, district)
        existing_coords = supabase.table("district_coordinates").select("district").eq("district", district).execute()
        if not existing_coords.data:
            coords = geocode_district(district)
            if coords:
                lat, lon = coords
                supabase.table("district_coordinates").insert({
                    "district": district, "latitude": lat, "longitude": lon,
                }).execute()

    supabase.table("participant_imports").update({
        "status": "confirmed",
        "confirmed_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", import_id).execute()

    return ImportActionResponse(message="Import confirmed.", id=import_id, row_count=len(rows))


@router.post("/imports/{import_id}/reject", response_model=ImportActionResponse)
async def reject_import(import_id: str, admin: dict = Depends(get_current_admin_user)):
    existing = supabase.table("participant_imports").select("id").eq("id", import_id).execute()
    if not existing.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found.")

    supabase.table("participant_imports").update({"status": "rejected"}).eq("id", import_id).execute()
    return ImportActionResponse(message="Import rejected.", id=import_id)


@router.delete("/imports/{import_id}", response_model=ImportActionResponse)
async def delete_import(
    import_id: str,
    confirm_cascade: bool = False,
    admin: dict = Depends(get_current_admin_user),
):
    """
    Deletes an import record.

    If the import was already confirmed, its rows were inserted into
    `participants`. Deleting the import record alone would leave those rows
    behind with a dangling source_import_id, so confirmed imports require
    the caller to pass ?confirm_cascade=true — this also deletes every
    participant row tied to this import. This double-confirmation exists
    because it changes numbers on the public-facing donor dashboard
    immediately and can't be undone.
    """
    existing = supabase.table("participant_imports").select("id, status").eq("id", import_id).execute()
    if not existing.data:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Import not found.")

    record = existing.data[0]
    deleted_participant_count = 0

    if record["status"] == "confirmed":
        participants_result = (
            supabase.table("participants").select("id").eq("source_import_id", import_id).execute()
        )
        deleted_participant_count = len(participants_result.data or [])

        if not confirm_cascade:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=(
                    f"This import is confirmed and linked to {deleted_participant_count} "
                    "participant record(s). Deleting it will also permanently remove those "
                    "records from the live dataset, which affects the public donor dashboard "
                    "immediately. Resend with confirm_cascade=true to proceed."
                ),
            )

        if deleted_participant_count > 0:
            supabase.table("participants").delete().eq("source_import_id", import_id).execute()

    supabase.table("participant_imports").delete().eq("id", import_id).execute()

    message = (
        f"Import deleted along with {deleted_participant_count} participant record(s)."
        if deleted_participant_count
        else "Import deleted."
    )
    return ImportActionResponse(message=message, id=import_id, row_count=deleted_participant_count)
