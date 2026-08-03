import csv
import io

import pypdf
from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.services.image_text_extractor import extract_text_from_image


def extract_pdf_text(file_bytes: bytes) -> str:
    reader = pypdf.PdfReader(io.BytesIO(file_bytes))
    text_parts = []

    for page in reader.pages:
        text_parts.append(page.extract_text() or "")

        for image in page.images:
            try:
                image_text = extract_text_from_image(image.data)
                if image_text:
                    text_parts.append(f"[Text found in embedded image '{image.name}']: {image_text}")
            except Exception as e:
                text_parts.append(f"[Could not read embedded image '{image.name}': {e}]")

    return "\n".join(text_parts)


def extract_docx_text(file_bytes: bytes) -> str:
    doc = DocxDocument(io.BytesIO(file_bytes))
    text_parts = [p.text for p in doc.paragraphs]

    for rel in doc.part.rels.values():
        if "image" in rel.target_ref:
            try:
                image_bytes = rel.target_part.blob
                image_text = extract_text_from_image(image_bytes)
                if image_text:
                    text_parts.append(f"[Text found in embedded image]: {image_text}")
            except Exception as e:
                text_parts.append(f"[Could not read an embedded image: {e}]")

    return "\n".join(text_parts)


def extract_xlsx_text(file_bytes: bytes) -> str:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"--- Sheet: {sheet.title} ---")
        for row in sheet.iter_rows(values_only=True):
            row_text = ", ".join(str(cell) if cell is not None else "" for cell in row)
            lines.append(row_text)
    return "\n".join(lines)


def extract_csv_text(file_bytes: bytes) -> str:
    text = file_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines = [", ".join(row) for row in reader]
    return "\n".join(lines)


def extract_text(filename: str, file_bytes: bytes) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "pdf":
        return extract_pdf_text(file_bytes)
    elif ext == "docx":
        return extract_docx_text(file_bytes)
    elif ext == "xlsx":
        return extract_xlsx_text(file_bytes)
    elif ext == "csv":
        return extract_csv_text(file_bytes)
    else:
        raise ValueError(f"Unsupported file type: {ext}")